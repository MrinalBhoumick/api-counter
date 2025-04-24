import streamlit as st
import json
import re
from collections import defaultdict
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
import requests
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
import matplotlib.pyplot as plt
import plotly.express as px

st.set_page_config(page_title="API Explorer", layout="wide")
st.title("🔍 API Structure Explorer & Load Tester")
st.markdown("Upload a **Postman Collection** or **OpenAPI/Swagger JSON** file to explore APIs, view env variables, and test performance.")

def extract_variables(text):
    return re.findall(r"{{(.*?)}}", text) if isinstance(text, str) else []

def parse_postman(data):
    folder_api_map = defaultdict(list)
    all_variables = set()
    total = 0

    def traverse(items, path=""):
        nonlocal total
        for item in items:
            name = item.get("name", "Unnamed")
            if 'request' in item:
                request = item['request']
                api_info = {"name": name}
                variables = set()
                url = request.get("url", {}).get("raw", "")
                variables.update(extract_variables(url))
                api_info["path"] = url
                for h in request.get("header", []):
                    variables.update(extract_variables(h.get("value", "")))
                body = request.get("body", {})
                if "raw" in body:
                    raw_body = body.get("raw", "")
                    variables.update(extract_variables(raw_body))
                    api_info["body"] = raw_body
                api_info["variables"] = sorted(list(variables))
                all_variables.update(variables)
                folder_api_map[path].append(api_info)
                total += 1
            elif 'item' in item:
                sub_path = f"{path}/{item['name']}" if path else item['name']
                traverse(item['item'], sub_path)

    traverse(data.get("item", []))
    return folder_api_map, total, all_variables

def parse_openapi(data):
    folder_api_map = defaultdict(list)
    all_variables = set()
    total = 0

    paths = data.get("paths", {})
    for endpoint, methods in paths.items():
        for method, details in methods.items():
            tag = details.get("tags", ["Untagged"])[0]
            summary = details.get("summary", f"{method.upper()} {endpoint}")
            api_info = {
                "name": f"{method.upper()} {endpoint} — {summary}",
                "path": endpoint,
                "variables": extract_variables(endpoint)
            }
            all_variables.update(api_info["variables"])
            request_body = details.get("requestBody", {})
            content = request_body.get("content", {})
            if "application/json" in content:
                example = content["application/json"].get("example")
                if example:
                    api_info["body"] = json.dumps(example, indent=2)
            folder_api_map[tag].append(api_info)
            total += 1

    return folder_api_map, total, all_variables

def load_test(api_urls, num_requests=10000, max_workers=100):
    results = []
    latencies = []
    statuses = []
    progress_bar = st.progress(0)

    def make_request(url):
        try:
            start = time.time()
            response = requests.get(url)
            latency = round((time.time() - start) * 1000, 2)
            latencies.append(latency)
            statuses.append(response.status_code)
            return (url, response.status_code, latency)
        except Exception as e:
            latencies.append(None)
            statuses.append("ERROR")
            return (url, "ERROR", str(e))

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = [executor.submit(make_request, url) for url in api_urls * (num_requests // len(api_urls))]
        completed = 0
        total = len(futures)
        for f in as_completed(futures):
            results.append(f.result())
            completed += 1
            progress_bar.progress(int((completed / total) * 100))

    return results, latencies, statuses

def plot_results(latencies, statuses):
    st.subheader("📊 Load Test Analysis")
    
    # Plot Latency Distribution
    fig = plt.figure(figsize=(10, 5))
    plt.hist(latencies, bins=50, color='skyblue', edgecolor='black')
    plt.title("Latency Distribution (ms)")
    plt.xlabel("Latency (ms)")
    plt.ylabel("Frequency")
    st.pyplot(fig)

    # Plot Success Rate
    success_count = statuses.count(200)
    error_count = len(statuses) - success_count
    success_rate = (success_count / len(statuses)) * 100
    error_rate = 100 - success_rate
    
    # Pie Chart for Success/Failure Rate
    fig = plt.figure(figsize=(7, 7))
    plt.pie([success_rate, error_rate], labels=["Success", "Error"], autopct='%1.1f%%', colors=["green", "red"], startangle=90)
    plt.title(f"Success Rate: {success_rate:.2f}%")
    st.pyplot(fig)

    # Plot Latency vs Status Code
    latency_df = pd.DataFrame({"Latency": latencies, "Status": statuses})
    fig = px.box(latency_df, x="Status", y="Latency", title="Latency by Status Code", points="all")
    st.plotly_chart(fig)

uploaded_file = st.file_uploader("📤 Upload your Postman/OpenAPI JSON file", type="json")

if uploaded_file:
    try:
        data = json.load(uploaded_file)
        folder_api_map = defaultdict(list)
        total = 0
        all_variables = set()

        if "item" in data:
            folder_api_map, total, all_variables = parse_postman(data)
            st.subheader("📂 Postman Collection: Folder-wise APIs")
        elif "paths" in data:
            folder_api_map, total, all_variables = parse_openapi(data)
            st.subheader("📂 OpenAPI/Swagger: Tag-wise APIs")
        else:
            st.error("❌ Unsupported format. Please upload a valid Postman or OpenAPI JSON file.")

        search_term = st.text_input("🔍 Search for API (by name or method)")
        filtered_apis = defaultdict(list)
        for folder, apis in folder_api_map.items():
            for api in apis:
                if search_term.lower() in api["name"].lower():
                    filtered_apis[folder].append(api)

        for folder, apis in filtered_apis.items():
            with st.expander(f"📁 {folder} — {len(apis)} APIs"):
                for api in apis:
                    st.markdown(f"**• {api['name']}**")
                    if api.get("path"):
                        st.markdown(f"  - 🌐 Path: `{api['path']}`")
                    if api.get("variables"):
                        st.markdown(f"  - 🔑 Env Vars: `{', '.join(api['variables'])}`")
                    if api.get("body"):
                        st.markdown("  - 📦 Request Body:")
                        st.code(api["body"], language="json")

        st.success(f"✅ Total API requests found: {total}")
        if all_variables:
            st.info(f"🌐 Unique environment variables used: `{', '.join(sorted(all_variables))}`")

        export_data = []
        for folder, apis in filtered_apis.items():
            for api in apis:
                export_data.append({
                    "Folder/Tag": folder,
                    "API Name": api["name"],
                    "API Path": api.get("path", ""),
                    "Env Variables": ", ".join(api.get("variables", [])),
                    "Request Body": api.get("body", ""),
                })

        if export_data:
            df = pd.DataFrame(export_data)
            buffer = BytesIO()
            wb = Workbook()
            ws = wb.active
            ws.title = "API Details"
            for row in dataframe_to_rows(df, index=False, header=True):
                ws.append(row)
            for cell in ws[1]:
                cell.font = Font(bold=True)
            for col in ws.columns:
                max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                ws.column_dimensions[col[0].column_letter].width = max_length + 2
            wb.save(buffer)
            buffer.seek(0)
            st.download_button(
                label="📥 Export as Excel",
                data=buffer,
                file_name="api_details.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        if st.button("🚀 Run Load Test (10k requests)"):
            st.info("Running load test... Please wait.")
            test_urls = [api["path"] for folder in filtered_apis for api in filtered_apis[folder] if api.get("path", "").startswith("http")]
            if not test_urls:
                st.warning("No valid full URLs found for testing.")
            else:
                results, latencies, statuses = load_test(test_urls, num_requests=10000)
                df_results = pd.DataFrame(results, columns=["URL", "Status", "Latency (ms)"])
                st.dataframe(df_results.head(100))
                buffer = BytesIO()
                wb = Workbook()
                ws = wb.active
                ws.title = "Load Test Results"
                for row in dataframe_to_rows(df_results, index=False, header=True):
                    ws.append(row)
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                for col in ws.columns:
                    max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                    ws.column_dimensions[col[0].column_letter].width = max_length + 2
                wb.save(buffer)
                buffer.seek(0)
                st.download_button(
                    label="📥 Export Load Test Results",
                    data=buffer,
                    file_name="load_test_results.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                st.success("✅ Load test completed!")
                plot_results(latencies, statuses)

    except Exception as e:
        st.error(f"❌ Error parsing file: {str(e)}")
